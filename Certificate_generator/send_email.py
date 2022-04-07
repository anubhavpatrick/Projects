'''
This module sends emails with attachments to the participants
Reference - https://developers.google.com/gmail/api/quickstart/python
'''

from __future__ import print_function

import os.path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from email.mime.text import MIMEText
import mimetypes
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
import base64
import openpyxl
import helpers
import config

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def aunthentication():
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'client_secrets.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

def prepare_email(event_name,result_file="results.xlsx"):
    """Prepares and send email with attachment to the participants 
    """
    creds = aunthentication()

    #formatting event name
    event_name_formatted = event_name.strip().lower().replace(' ','_').replace('.','_')

    # loading the details.xlsx workbook
    # and grabbing the active sheet
    obj = openpyxl.load_workbook(result_file)
    sheet = obj.active

    # iterate over rows in excel sheet
    for i in range(2, sheet.max_row+1):
        # grab values from excel sheet
        try:
            get_name = sheet.cell(row = i ,column = 4)
            participant_name = get_name.value.title()

            #Abbreviate name if |name| > 26
            if len(participant_name) > 26:
                participant_name = helpers.abbreviated_name(participant_name)

            get_section = sheet.cell(row = i ,column = 2)
            section = get_section.value.lower()

            get_email = sheet.cell(row = i ,column = 5)
            participant_email = get_email.value

            get_roll_no = sheet.cell(row = i ,column = 3)
            roll_no = get_roll_no.value

            participant_name_formatted = participant_name.strip().lower().replace(' ','_').replace('.','_')
            attachment_file_path= f"{event_name_formatted}/sec_{section}/{section}_{roll_no}_{participant_name_formatted}.pdf"

            # Call the Gmail API
            service = build('gmail', 'v1', credentials=creds)

            # get the email body and subject
            subject, body = create_email_subject_body(event_name, participant_name)

            #create message with attachment
            msg = create_message_with_attachment(config.DEFAULT_SENDER_EMAIL, participant_email, subject, body, attachment_file_path)
            send_message(service, 'me', msg)

        except HttpError as error:
            # TODO(developer) - Handle errors from gmail API.
            print(f'An error occurred: {error}')

def create_message(sender, to, subject, message_text):
    """Create a message for an email.

    Args:
    sender: Email address of the sender.
    to: Email address of the receiver.
    subject: The subject of the email message.
    message_text: The text of the email message.

    Returns:
    An object containing a base64url encoded email object.
    """
    message = MIMEText(message_text)
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject
    return {'raw': base64.urlsafe_b64encode(message.as_string().encode()).decode()}

def create_message_with_attachment(
    sender, to, subject, message_text, file):
    """Create a message for an email.

    Args:
    sender: Email address of the sender.
    to: Email address of the receiver.
    subject: The subject of the email message.
    message_text: The text of the email message.
    file: The path to the file to be attached.

    Returns:
    An object containing a base64url encoded email object.
    """
    message = MIMEMultipart()
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject

    msg = MIMEText(message_text)
    message.attach(msg)

    content_type, encoding = mimetypes.guess_type(file)

    if content_type is None or encoding is not None:
        content_type = 'application/octet-stream'
    main_type, sub_type = content_type.split('/', 1)
    if main_type == 'text':
        fp = open(file, 'rb')
        msg = MIMEText(fp.read(), _subtype=sub_type)
        fp.close()
    elif main_type == 'image':
        fp = open(file, 'rb')
        msg = MIMEImage(fp.read(), _subtype=sub_type)
        fp.close()
    else:
        fp = open(file, 'rb')
        msg = MIMEBase(main_type, sub_type)
        msg.set_payload(fp.read())
        fp.close()
    filename = os.path.basename(file)
    msg.add_header('Content-Disposition', 'attachment', filename=filename)
    message.attach(msg)

    return {'raw': base64.urlsafe_b64encode(message.as_string().encode()).decode()}

def send_message(service, user_id, message):
    """Send an email message.

    Args:
    service: Authorized Gmail API service instance.
    user_id: User's email address. The special value "me"
    can be used to indicate the authenticated user.
    message: Message to be sent.

    Returns:
    Sent Message.
    """
    try:
        message = (service.users().messages().send(userId=user_id, body=message)
                    .execute())
        print('Message Id: %s' % message['id'])
        return message
    except HttpError as error:
        print('An error occurred: %s' % error)

def create_email_subject_body(event_name:str, participant_name, test_mode=False):
    '''Create subject and body of the email to be sent'''
    subject = f"Certificate of competency in {event_name}"
    if test_mode:
        body = "This is a test email.\n\n"
    else:
        body = ""
    body +=  f"Dear {participant_name},\n\n" +\
            f"Congratulations! You have successfully completed the " + f"{event_name.strip()}" +" event. \n\nPlease find your attached certificate of competency.\n\n" +\
            f"Best Wishes,\n" +\
            f"Team CPP"
    return subject, body

'''if __name__ == '__main__':
    #send_email()
    prepare_email("CODEWIZ 1.0", "results.xlsx")'''
