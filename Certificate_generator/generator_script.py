# import the necessary libraries</pre>

import cv2 as cv
import openpyxl
from PIL import Image
import os
import datetime

def generate_certificate(result_file, template_path, session, event_name:str):
    
    # pad event name with spaces
    event_name_padded = event_name.center(20, ' ')

    # pad session with spaces
    session_padded = session.center(15, ' ')

    # Setting the font size and font
    # colour
    font_size = 3
    font_color = (255, 255, 255)

    # Coordinates on the certificate where
    # will be printing the name (set
    # according to your own template)
    coordinate_y_adjustment = 15
    coordinate_x_adjustment = 7

    #settings for event name to be put on the certificate
    text_x_event_name = 405
    text_y_event_name = 505
    font_event_name = cv.FONT_HERSHEY_COMPLEX_SMALL
    font_size_event_name = 1
    font_thickness_event_name = 2

    #settings for session to be put on the certificate
    text_x_session = 840
    text_y_session = 455
    font_session = cv.FONT_HERSHEY_COMPLEX_SMALL
    font_size_session = 1
    font_thickness_session = 2

    #Get current date in DD-MM-YYYY format
    current_date = datetime.datetime.now()
    certificate_issue_date = 'Issue Date: ' + current_date.strftime("%d-%m-%Y")
    #settings for issue_date to be put on the certificate
    font_date = cv.FONT_HERSHEY_PLAIN
    font_size_date = 1
    font_thickness_date = 1
     # get the size of the name to be printed
    date_text_size = cv.getTextSize(certificate_issue_date, font_date, font_size_date, font_thickness_date)[0]

    # loading the details.xlsx workbook
    # and grabbing the active sheet
    obj = openpyxl.load_workbook(result_file)
    sheet = obj.active

    #store dir listing for creating new directories (if necessary)
    dir_list = os.listdir()

    # excel sheet
    for i in range(2, sheet.max_row+1):
    
        # grabs the row=i and column=1 cell
        # that contains the name value of that
        # cell is stored in the variable certi_name
        try:
            get_name = sheet.cell(row = i ,column = 4)
            certi_name = get_name.value.title()

            get_section = sheet.cell(row = i ,column = 2)
            section = get_section.value.lower()
            # Output Paths
            event_name_formatted = event_name.strip().lower().replace(' ','_').replace('.','_')
            output_path = event_name_formatted+'/sec_' + section.lower() + '/'

            get_roll_num = sheet.cell(row = i ,column = 3)
            roll_num = int(get_roll_num.value)
        except:
            #if some entries are blank then skip entire row
            continue

        #if directory for current section is not created then create if
        d = event_name_formatted+'/sec_' + section
        if d not in dir_list:
            os.makedirs(d,exist_ok=True)
            dir_list = os.listdir()

        # read the certificate template
        img = cv.imread(template_path)

        # choose the font from opencv
        font = cv.FONT_HERSHEY_SCRIPT_SIMPLEX

        # get the size of the name to be
        # printed
        text_size = cv.getTextSize(certi_name, font, font_size, 5)[0]

        # get the (x,y) coordinates where the
        # name is to written on the template
        # The function cv.putText accepts only
        # integer arguments so convert it into 'int'.
        text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
        text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
        text_x = int(text_x)
        text_y = int(text_y)+10
        cv.putText(img, certi_name,
                (text_x, text_y),
                font,
                font_size,
                font_color, 5)
        
        #displaying event name on the certificate
        cv.putText(img, event_name_padded,
                (text_x_event_name, text_y_event_name),
                font_event_name,
                font_size_event_name,
                font_color, font_thickness_event_name)
        
        #displaying session on the certificate
        cv.putText(img, session_padded,
                (text_x_session, text_y_session),
                font_session,
                font_size_session,
                font_color, font_thickness_session)
        
        #displaying issue date on the certificate
        text_x_date = 10
        text_y_date = (img.shape[0] - (date_text_size[1]+2))
        cv.putText(img, certificate_issue_date,
                (text_x_date, text_y_date),
                font_date,
                font_size_date,
                font_color, font_thickness_date)

        # Output path along with the name of the
        # certificate generated
        certi_name = certi_name.lower().replace(' ','_')
        certi_path = output_path + section + '_' + str(roll_num) + '_' + certi_name

        # Save the certificate
        cv.imwrite(certi_path+'.png', img)

        image_1 = Image.open(certi_path + '.png')
        im_1 = image_1.convert('RGB')
        im_1.save(certi_path +'.pdf')

        open('.gitignore', 'a').write(certi_path + '/\n')
